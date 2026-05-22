#!/usr/bin/env python3
"""
update_dividend_calendar.py
─────────────────────────────────────────────────────────────────────────
Parses the manually-curated `dividend.xlsx` into
`output/dividend-calendar.json` — the forward-looking dividend events
calendar the iOS app's Calendar tab consumes.

Each row in the xlsx is one declared dividend payout:
    ISIN | security code (= our tickerNumber) | ticker | cum date | ex date | record date | payment date

The output is a flat chronological list, newest cum-date first. The app
groups by ex-date for display. We also cross-reference each row against
our `boursa-kuwait-stocks.json` to attach the canonical English + Arabic
company name so the calendar view doesn't need a separate join at
render time.

Run cadence: ad-hoc. Whenever Ali drops a refreshed xlsx into `data/`,
re-run this script + commit + deploy. The xlsx is the source of truth
(curated by hand from Boursa Kuwait's corporate-actions announcements);
this script is just a serializer.

Usage:
    python3 update_dividend_calendar.py
    python3 update_dividend_calendar.py --xlsx dividend.xlsx --out output/dividend-calendar.json
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = SCRIPT_DIR / "dividend.xlsx"
DEFAULT_OUT = SCRIPT_DIR / "output" / "dividend-calendar.json"
STOCKS_FILE = SCRIPT_DIR / "boursa-kuwait-stocks.json"


def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


# Column headers as they appear in the source xlsx (with their typos).
# Keep this list synced with what Ali maintains.
COL_TICKER = 2  # zero-indexed
COL_CUM = 3
COL_EX = 4
COL_RECORD = 5
COL_PAY = 6


def iso_date(value) -> str | None:
    """Convert a cell value (datetime or string) to ISO YYYY-MM-DD."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    # Sometimes Excel hands back a date as a string already; pass through.
    if isinstance(value, str):
        return value.strip() or None
    return str(value)


def load_stock_lookup() -> dict[str, dict]:
    """Map ticker → {nameEn, nameAr, tickerNumber, sector} so the calendar
    can render the company name + sector without a runtime join."""
    if not STOCKS_FILE.exists():
        return {}
    with open(STOCKS_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    return {
        s["ticker"]: {
            "nameEn": s.get("nameEn"),
            "nameAr": s.get("nameAr"),
            "tickerNumber": s.get("tickerNumber"),
            "sector": s.get("sector"),
        }
        for s in data["stocks"]
    }


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Path to the source xlsx")
    ap.add_argument("--out", default=str(DEFAULT_OUT), help="Output JSON path")
    return ap.parse_args()


def main() -> int:
    args = parse_args()
    xlsx_path = Path(args.xlsx)
    out_path = Path(args.out)
    if not xlsx_path.exists():
        log(f"ERROR: xlsx not found at {xlsx_path}")
        return 1

    stock_lookup = load_stock_lookup()
    if not stock_lookup:
        log(f"WARN: stock lookup empty (couldn't read {STOCKS_FILE.name}); names will be null")

    log(f"Reading {xlsx_path.name}…")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    events: list[dict] = []
    seen_tickers: set[str] = set()
    rows_processed = 0
    rows_skipped = 0

    # Skip the header row
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row is None or all(v is None for v in row):
            continue
        rows_processed += 1
        try:
            ticker_cell = row[COL_TICKER]
            if ticker_cell is None:
                rows_skipped += 1
                continue
            ticker = str(ticker_cell).strip().upper()
            cum_date = iso_date(row[COL_CUM])
            ex_date = iso_date(row[COL_EX])
            record_date = iso_date(row[COL_RECORD])
            payment_date = iso_date(row[COL_PAY])
        except IndexError:
            rows_skipped += 1
            continue

        meta = stock_lookup.get(ticker, {})
        events.append({
            "ticker": ticker,
            "nameEn": meta.get("nameEn"),
            "nameAr": meta.get("nameAr"),
            "tickerNumber": meta.get("tickerNumber"),
            "sector": meta.get("sector"),
            "cumDate": cum_date,
            "exDate": ex_date,
            "recordDate": record_date,
            "paymentDate": payment_date,
        })
        seen_tickers.add(ticker)

    # Sort by ex-date descending (newest first) — what the app groups by
    events.sort(key=lambda e: (e.get("exDate") or "", e["ticker"]), reverse=True)

    unmatched = sorted(t for t in seen_tickers if t not in stock_lookup)
    log(f"Parsed {rows_processed} rows → {len(events)} events ({len(seen_tickers)} unique tickers)")
    if rows_skipped:
        log(f"  Skipped {rows_skipped} blank/invalid rows")
    if unmatched:
        log(f"  {len(unmatched)} tickers not in our stock list: {unmatched[:10]}…" if len(unmatched) > 10 else f"  Unmatched tickers: {unmatched}")

    output = {
        "_meta": {
            "generatedAt": datetime.utcnow().isoformat() + "Z",
            "source": "manual xlsx maintained by Ali",
            "eventCount": len(events),
        },
        "events": events,
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    log(f"Wrote {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
